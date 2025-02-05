import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import datetime
import os
import sys
import logging
from urllib.parse import quote
from openpyxl.utils import get_column_letter

def fetch_twse_data(date):
    url = f"https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?date={date}&type=ALLBUT0999&response=html"
    response = requests.get(url)
    if response.status_code != 200:
        print("無法從 TWSE 抓取資料")
        return None

    soup = BeautifulSoup(response.content, 'html.parser')
    if not soup.find_all('table'):
        print(f"{date} 無交易資料，尋找前一個交易日的資料...")
        previous_date = (datetime.datetime.strptime(date, "%Y%m%d") - datetime.timedelta(days=1)).strftime("%Y%m%d")
        return fetch_twse_data(previous_date)

    data = {}
    tables = soup.find_all('table')
    table_names = [
        "價格指數(臺灣證券交易所)",
        "價格指數(跨市場)",
        "價格指數(臺灣指數公司)",
        "報酬指數(臺灣證券交易所)",
        "報酬指數(跨市場)",
        "報酬指數(臺灣指數公司)",
        "大盤統計資訊",
        "漲跌證券數合計",
        "每日收盤行情(全部(不含權證、牛熊證))",
    ]

    for i, table in enumerate(tables):
        if i < len(table_names):
            table_name = table_names[i]
            headers = [th.get_text(strip=True) for th in table.find_all('th')]
            rows = []
            for tr in table.find_all('tr')[1:]:
                rows.append([td.get_text(strip=True) for td in tr.find_all('td')])
            data[table_name] = {"headers": headers, "rows": rows}

    return data

def fetch_tpex_data():
    current_date = datetime.datetime.now().strftime("%Y/%m/%d")
    encoded_date = quote(current_date, safe='')
    url = f"https://www.tpex.org.tw/www/zh-tw/afterTrading/otc?date={encoded_date}&type=EW&response=html&order=0&sort=asc"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            table = soup.find('table')
            headers = [th.get_text(strip=True) for th in table.find_all('th')]
            rows = []
            for tr in table.find_all('tr')[1:]:
                rows.append([td.get_text(strip=True) for td in tr.find_all('td')])
            return {"headers": headers, "rows": rows}
        else:
            print(f"请求失败，状态码：{response.status_code}")
            return None
    except Exception as e:
        print(f"发生错误：{e}")
        return None

def save_to_excel(twse_data, tpex_data, output_file):
    wb = Workbook()
    
    if twse_data:
        for key, value in twse_data.items():
            ws = wb.create_sheet(title=key)
            headers = value["headers"]
            rows = value["rows"]

            # 寫入標題（不變）
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            # **僅對"每日收盤行情(全部(不含權證、牛熊證))"的數據右移**
            if key == "每日收盤行情(全部(不含權證、牛熊證))":
                for row_num, row_data in enumerate(rows, 2):
                    ws.cell(row=row_num, column=1, value="")  # 第一欄變空白
                    for col_num, cell_data in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num + 1, value=cell_data)  # 向右移動
            else:
                # 其他表格照原本方式寫入
                for row_num, row_data in enumerate(rows, 2):
                    for col_num, cell_data in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=cell_data)

    if tpex_data:
        ws = wb.create_sheet(title="上櫃股票每日收盤行情(不含定價)")
        headers = tpex_data["headers"]
        rows = tpex_data["rows"]

        # 寫入標題（不變）
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # 直接寫入數據（不做右移）
        for row_num, row_data in enumerate(rows, 2):
            for col_num, cell_data in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_data)

    # **調整欄寬**
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 取得欄位名稱
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(output_file)


     # 添加 VBA 宏
    if 'VBA' not in wb.sheetnames:
        ws = wb.create_sheet(title='VBA')
        ws['A1'] = 'Sub Auto_Open()'
        ws['A2'] = '    Call UpdateData'
        ws['A3'] = 'End Sub'
        ws['A4'] = 'Sub UpdateData()'
        ws['A5'] = '    Dim objShell As Object'
        ws['A6'] = '    Set objShell = CreateObject("WScript.Shell")'
        ws['A7'] = '    objShell.Run "python d:/code/python/Push_shares/push_html.py"'
        ws['A8'] = 'End Sub'

    wb.save(output_file)
    
def main():
    # 檢查 Excel 檔案是否存在
    output_file = 'twse_data.xlsx'
    if os.path.exists(output_file):
        # 設定初始日期為開啟 Excel 檔案的當天日期
        date = datetime.datetime.now().strftime("%Y%m%d")
    else:
        # 設定初始日期為當天日期
        date = datetime.datetime.now().strftime("%Y%m%d")
    
    # 從 TWSE 網站抓取資料
    twse_data = fetch_twse_data(date)
    # 從 TPEx 網站抓取資料
    tpex_data = fetch_tpex_data()
    
    if twse_data or tpex_data:
        # 保存資料到 Excel 檔案
        save_to_excel(twse_data, tpex_data, output_file)
        print(f"資料已保存到 {output_file}")

if __name__ == "__main__":
    try:
        main()
        # 自動打開文件（Windows系統）
        if os.name == 'nt':
            os.startfile(os.path.abspath("twse_data.xlsx"))
        logging.info("操作成功完成")
    except Exception as e:
        logging.critical(f"系統錯誤: {str(e)}")
        sys.exit(2)
