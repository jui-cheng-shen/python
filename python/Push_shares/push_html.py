import requests
import pandas as pd
from bs4 import BeautifulSoup
from io import StringIO
import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

def fetch_twse_data():
    url = "https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?date=20250117&type=ALLBUT0999&response=html"
    
    # 使用 Selenium 來抓取動態生成的內容
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    driver.get(url)
    
    try:
        # 等待表格加載完成
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "table"))
        )
        soup = BeautifulSoup(driver.page_source, "html.parser")
        table = soup.find("table")
        return pd.read_html(StringIO(str(table)))[0]
    finally:
        driver.quit()

def fetch_tpex_data():
    url = "https://www.tpex.org.tw/www/zh-tw/afterTrading/otc?date=2025%2F01%2F17&type=EW&response=html&order=0&sort=asc"
    response = requests.get(url)
    if response.status_code == 200:
        print("TPEx 網站資料成功抓取")
        soup = BeautifulSoup(response.text, "html.parser")
        table = soup.find("table")
        return pd.read_html(StringIO(str(table)))[0]
    else:
        print("無法抓取 TPEx 網站資料")
        return None

def save_to_excel(tpex_data, twse_data, output_file):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        if tpex_data is not None:
            # 展平多重索引
            tpex_data.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in tpex_data.columns.values]
            tpex_data.to_excel(writer, sheet_name='股票(上櫃)', index=False)
        if twse_data is not None:
            # 展平多重索引
            twse_data.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in twse_data.columns.values]
            twse_data.to_excel(writer, sheet_name='股票(上市)', index=False)

    # 調整每個儲存格的大小
    wb = openpyxl.load_workbook(output_file, keep_vba=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # 添加 VBA 宏
    if 'VBA' not in wb.sheetnames:
        ws = wb.create_sheet(title='VBA')
        ws['A1'] = 'Sub Auto_Open()'
        ws['A2'] = '    Call UpdateData'
        ws['A3'] = 'End Sub'
        ws['A4'] = 'Sub UpdateData()'
        ws['A5'] = '    Dim objShell As Object'
        ws['A6'] = '    Set objShell = CreateObject("WScript.Shell")'
        ws['A7'] = '    objShell.Run "python d:/code/python/Push_shares/push_html.py"'
        ws['A8'] = 'End Sub'

    wb.save(output_file)

def main():
    # 從 TWSE 網站抓取資料
    twse_data = fetch_twse_data()
    if twse_data is not None:
        print("TWSE 資料：")
        print(twse_data)

    # 從 TPEx 網站抓取資料
    tpex_data = fetch_tpex_data()
    if tpex_data is not None:
        print("TPEx 資料：")
        print(tpex_data)

    # 儲存資料到 Excel 檔案
    output_file = "trading_data_new.xlsm"  # 使用不同的檔案名稱
    save_to_excel(tpex_data, twse_data, output_file)

if __name__ == "__main__":
    main()