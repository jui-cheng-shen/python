import requests
import pandas as pd
from bs4 import BeautifulSoup
from io import StringIO
import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
import logging

# 配置日志记录
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

def get_formatted_date():
    """获取当前日期并格式化为所需的字符串格式"""
    today = datetime.now()
    return {
        'twse_date': today.strftime("%Y%m%d"),
        'tpex_date': today.strftime("%Y/%m/%d")
    }

def fetch_twse_data(date):
    """抓取台湾证券交易所数据"""
    url = f"https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?date={date}&type=ALLBUT0999&response=html"
    
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    
    try:
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=options
        )
        
        driver.get(url)
        
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-b"))
        )
        
        soup = BeautifulSoup(driver.page_source, "html.parser")
        table = soup.find("table", class_="table-b")
        
        if not table:
            logging.error("TWSE表格未找到")
            return None
            
        df = pd.read_html(StringIO(str(table)))[0]
        
        # 处理多重索引列
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = ['_'.join(col).strip() for col in df.columns.values]
        
        return df
        
    except Exception as e:
        logging.error(f"抓取TWSE数据失败: {str(e)}")
        return None
    finally:
        if 'driver' in locals():
            driver.quit()

def fetch_tpex_data(date):
    """抓取柜买中心数据"""
    encoded_date = date.replace("/", "%2F")
    url = f"https://www.tpex.org.tw/www/zh-tw/afterTrading/otc?date={encoded_date}&type=EW&response=html"
    
    try:
        response = requests.get(url, timeout=15)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, "html.parser")
        table = soup.find("table", {"class": "table-b"})
        
        if not table:
            logging.error("TPEx表格未找到")
            return None
            
        df = pd.read_html(StringIO(str(table)))[0]
        
        # 清理列名
        df.columns = [col[1] if isinstance(col, tuple) else col for col in df.columns]
        
        return df
        
    except Exception as e:
        logging.error(f"抓取TPEx数据失败: {str(e)}")
        return None

def optimize_column_widths(ws):
    """优化Excel列宽"""
    for col in ws.columns:
        max_length = max((
            len(str(cell.value)) 
            for cell in col 
            if cell.value is not None
        ), default=10)
        adjusted_width = (max_length + 2) * 1.2
        column_letter = get_column_letter(col[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width

def save_to_excel(tpex_data, twse_data, output_file):
    """保存数据到Excel文件"""
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            if tpex_data is not None:
                tpex_data.to_excel(
                    writer, 
                    sheet_name='股票(上櫃)', 
                    index=False,
                    header=[col.replace('\n', ' ') for col in tpex_data.columns]
                )
                
            if twse_data is not None:
                twse_data.to_excel(
                    writer, 
                    sheet_name='股票(上市)', 
                    index=False,
                    header=[col.replace('\n', ' ') for col in twse_data.columns]
                )

        # 优化列宽
        wb = openpyxl.load_workbook(output_file)
        for sheet_name in wb.sheetnames:
            if sheet_name == 'VBA':
                continue
            optimize_column_widths(wb[sheet_name])

        wb.save(output_file)
        logging.info(f"文件已保存至 {output_file}")

    except Exception as e:
        logging.error(f"保存Excel文件失败: {str(e)}")
        raise

def main():
    dates = get_formatted_date()
    
    logging.info("开始抓取TWSE数据...")
    twse_data = fetch_twse_data(dates['twse_date'])
    
    logging.info("开始抓取TPEx数据...")
    tpex_data = fetch_tpex_data(dates['tpex_date'])
    
    output_file = f"trading_data_{datetime.now().strftime('%Y%m%d')}.xlsx"
    save_to_excel(tpex_data, twse_data, output_file)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logging.error(f"程序运行失败: {str(e)}")